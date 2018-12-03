'''
@date 29/11/2018
@author Minh Dinh

input: excel files
output: aiml files

'''
import os
import constant

import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def validate_cell(input_cell):
    if not input_cell or input_cell == '':
        return False
    return True


'''
    read excel files:
        + init list of xcel files
        + ierate through list of file:
            + only get xlsx files into a list            
'''
def get_xlsx_files(xlsx_folder):
    list_of_xcel_files = []
    for f in os.listdir(xlsx_folder):
        if f.endswith('xlsx') and f not in list_of_xcel_files:
            list_of_xcel_files.append(f)
    return list_of_xcel_files


'''
    read xcel:
        + dict init
        + iterate through sheet:
            + validate k,v
            + add k,v to dict
'''
def read_xlsx_file(xlsx_file_as_path):
    qa_as_dict = {}
    workbook = load_workbook(filename=os.path.join(xlsx_file_as_path))
    for each_sheet in workbook.sheetnames:
        worksheet = workbook[each_sheet]
        for row in range(1,worksheet.max_row):
            if validate_cell(str(worksheet.cell(column=1,row=row).value)) and validate_cell(str(worksheet.cell(column=2,row=row).value)):
                qa_as_dict[str(worksheet.cell(column=1,row=row).value)] = \
                str(worksheet.cell(column=2,row=row).value)
    return qa_as_dict


'''
    processing data as k:v
        + input: data as k:v
        + output: pattern : template
'''
# group questions that have the same answer 

# analyze question >> remove stopwords, etc....




'''
   create aiml files
'''
def create_and_fill_single_aiml_file(name, location,content_as_dict):
    with open(os.path.join(location,name+".aiml"),'w') as aiml_file:
        aiml_file.write('<?xml version="1.0" encoding="UTF-8"?>')
        aiml_file.write("\n")
        aiml_file.write('<aiml version="2.0">')
        aiml_file.write("\n")

        for k,v in content_as_dict.items():
            aiml_file.write("\t<category>") 
            aiml_file.write("\n")
            aiml_file.write("\t\t<pattern>")        
            aiml_file.write("\t\t\t"+k)        
            aiml_file.write("\t\t</pattern>")
            aiml_file.write("\n")
            aiml_file.write("\t\t<template>")                        
            aiml_file.write("\t\t\t"+v)        
            aiml_file.write("\t\t</template>")
            aiml_file.write("\n")
            aiml_file.write("\t</category>") 
            aiml_file.write("\n\n")

        aiml_file.write("</aiml>")
# create aiml header n footer

# write pattern : template

#============ TeST==============
list_of_xcel_files = get_xlsx_files(constant.INPUT_XCEL_FILES)

for each in list_of_xcel_files:
    content_as_dict = read_xlsx_file(os.path.join(constant.INPUT_XCEL_FILES,each))
    create_and_fill_single_aiml_file(each,constant.OUTPUT_AIML_FILES,content_as_dict)
    